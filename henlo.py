import os
import shutil
import pandas as pd
from openpyxl import load_workbook
from tkinter import filedialog, Tk

# Step 1: Function to initialize the master file if it doesn't exist
def initialize_master_file():
    master_path = 'Master_Form.xlsx'
    if not os.path.exists(master_path):
        # Create a new master file with default structure
        df = pd.DataFrame(columns=['Name', 'Gender', 'Grade', 'Section'])
        df.to_excel(master_path, index=False)
    return master_path

# Step 2: Function to set up output directories and copy template files
def setup_output_directories(output_folder):
    files_to_copy = ['SF1.xlsx', 'SF5A.xlsx', 'SF5B.xlsx']
    output_dirs = {}
    for file_name in files_to_copy:
        # Make subdirectory for each file type
        sub_dir = os.path.join(output_folder, os.path.splitext(file_name)[0])
        os.makedirs(sub_dir, exist_ok=True)
        # Copy file to output subdirectory
        dest_path = os.path.join(sub_dir, file_name)
        shutil.copy(file_name, dest_path)
        output_dirs[file_name] = dest_path
    return output_dirs

# Step 3: Function to read and process data from the master form
def read_and_sort_master_data(master_path):
    df = pd.read_excel(master_path)
    df = df.sort_values(by=['Name'])  # Sort by name alphabetically
    male_df = df[df['Gender'] == 'Male']
    female_df = df[df['Gender'] == 'Female']
    return male_df, female_df

# Step 4: Function to populate target files (SF1, SF5A, SF5B) with sorted data
def populate_files(output_dirs, male_df, female_df):
    config = {
        "SF1.xlsx": {"male_row_start": 5, "female_row_start": 15},
        "SF5A.xlsx": {"male_row_start": 7, "female_row_start": 17},
        "SF5B.xlsx": {"male_row_start": 8, "female_row_start": 18}
    }
    for file_name, path in output_dirs.items():
        wb = load_workbook(path)
        ws = wb.active
        cfg = config[file_name]
        
        # Populate males
        for i, (_, row) in enumerate(male_df.iterrows(), start=cfg["male_row_start"]):
            ws[f"A{i}"] = row["Name"]
            ws[f"B{i}"] = row["Grade"]
            ws[f"C{i}"] = row["Section"]
        
        # Populate females
        for i, (_, row) in enumerate(female_df.iterrows(), start=cfg["female_row_start"]):
            ws[f"A{i}"] = row["Name"]
            ws[f"B{i}"] = row["Grade"]
            ws[f"C{i}"] = row["Section"]
        
        wb.save(path)

# Main execution
def main():
    # Initialize master file if missing
    master_path = initialize_master_file()
    
    # Prompt user for output folder
    root = Tk()
    root.withdraw()
    output_folder = filedialog.askdirectory(title="Select Output Folder")
    if not output_folder:
        print("No output folder selected. Exiting.")
        return

    # Setup output directories and copy template files
    output_dirs = setup_output_directories(output_folder)
    
    # Read and process data from master form
    male_df, female_df = read_and_sort_master_data(master_path)
    
    # Populate each file in output directories with sorted data
    populate_files(output_dirs, male_df, female_df)

    print("Data processing and population complete. Files saved to respective subdirectories.")

# Run the main function
main()
